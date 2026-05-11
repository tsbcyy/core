# -*- coding: utf-8 -*-

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import sqlite3

from database import Database
from ocr_utils import extract_points_advanced

# 初始化Flask应用
app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production-2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}

# 初始化数据库
db = Database()

# 初始化Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    if user_id == '123456':
        return User(user_id)
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==================== 路由定义 ====================

@app.route('/')
def index():
    """首页 - 选择入口"""
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """管理员登录"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == '123456' and password == '123456':
            user = User(username)
            login_user(user)
            flash('✅ 登录成功！', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('❌ 账号或密码错误', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('已退出登录', 'info')
    return redirect(url_for('login'))

@app.route('/submit', methods=['GET', 'POST'])
def submit():
    """提交者页面"""
    if request.method == 'POST':
        try:
            student_id = request.form.get('student_id', '').strip()
            student_name = request.form.get('student_name', '').strip()
            
            # 验证学号姓名
            conn = sqlite3.connect('database.db')
            c = conn.cursor()
            c.execute("SELECT name, class FROM students WHERE id=?", (student_id,))
            row = c.fetchone()
            
            if not row:
                conn.close()
                flash('❌ 错误：学号不存在，请联系管理员', 'error')
                return redirect(url_for('submit'))
            
            if row[0] != student_name:
                conn.close()
                flash('❌ 错误：姓名与学号不匹配', 'error')
                return redirect(url_for('submit'))
            
            student_class = row[1] if row[1] else ''
            conn.close()
            
            # 处理图片上传
            if 'image' not in request.files:
                flash('❌ 请选择图片文件', 'error')
                return redirect(url_for('submit'))
            
            image = request.files['image']
            if image.filename == '':
                flash('❌ 未选择文件', 'error')
                return redirect(url_for('submit'))
            
            if not allowed_file(image.filename):
                flash('❌ 不支持的文件格式，请上传图片文件', 'error')
                return redirect(url_for('submit'))
            
            # 保存图片
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = secure_filename(f"{student_id}_{timestamp}_{image.filename}")
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image.save(image_path)
            
            # 获取用户输入的积分
            manual_points_str = request.form.get('manual_points', '').strip()
            user_manual_points = int(manual_points_str) if manual_points_str.isdigit() else None
            
            # 高级OCR识别
            ocr_result = extract_points_advanced(image_path, user_manual_points)
            
            if not ocr_result['success']:
                error_msg = ocr_result.get('error', 'OCR识别失败')
                db.add_submission(
                    student_id, student_name, student_class,
                    user_manual_points, None,
                    image_path, error_msg
                )
                flash(f'⚠️ 识别失败: {error_msg}', 'error')
                return render_template('submit_result.html',
                                     ocr_points=None,
                                     manual_points=user_manual_points,
                                     match=False,
                                     student_name=student_name,
                                     error_message=error_msg)
            
            ocr_points = ocr_result['points']
            
            if user_manual_points is not None and user_manual_points != ocr_points:
                error_msg = f"用户输入{user_manual_points}与系统识别{ocr_points}不一致"
                db.add_submission(
                    student_id, student_name, student_class,
                    user_manual_points, ocr_points,
                    image_path, error_msg
                )
                flash(f'⚠️ 您输入的积分({user_manual_points})与图片识别结果({ocr_points})不一致！已通知管理员审核。', 'warning')
                return render_template('submit_result.html',
                                     ocr_points=ocr_points,
                                     manual_points=user_manual_points,
                                     match=False,
                                     student_name=student_name,
                                     error_message=error_msg)
            else:
                final_points = user_manual_points if user_manual_points is not None else ocr_points
                db.add_submission(
                    student_id, student_name, student_class,
                    final_points, ocr_points,
                    image_path, None
                )
                flash(f'✅ 提交成功！系统识别积分为: {ocr_points}分', 'success')
                return render_template('submit_result.html',
                                     ocr_points=ocr_points,
                                     manual_points=user_manual_points,
                                     match=True,
                                     student_name=student_name,
                                     error_message=None)
        
        except Exception as e:
            flash(f'系统错误: {str(e)}', 'error')
            return redirect(url_for('submit'))
    
    return render_template('submit.html')

@app.route('/admin')
@login_required
def admin_dashboard():
    stats = db.get_statistics()
    pending_errors = db.get_pending_errors()
    return render_template('admin_dashboard.html', stats=stats, pending_errors=pending_errors)

@app.route('/admin/students', methods=['GET', 'POST'])
@login_required
def manage_students():
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add_single':
            student_id = request.form.get('student_id', '').strip()
            name = request.form.get('name', '').strip()
            class_name = request.form.get('class', '').strip()
            
            if student_id and name:
                db.add_student(student_id, name, class_name)
                flash(f'✅ 学生 {name}({student_id}) 添加成功', 'success')
            else:
                flash('❌ 学号和姓名不能为空', 'error')
        
        elif action == 'batch_import':
            data = request.form.get('batch_data', '').strip()
            lines = data.split('\n')
            success_count = 0
            error_count = 0
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                parts = line.split(',')
                if len(parts) >= 2:
                    student_id = parts[0].strip()
                    name = parts[1].strip()
                    class_name = parts[2].strip() if len(parts) > 2 else ''
                    db.add_student(student_id, name, class_name)
                    success_count += 1
                else:
                    error_count += 1
            
            flash(f'✅ 批量导入完成: 成功{success_count}条，失败{error_count}条', 'success')
        
        elif action == 'delete':
            student_id = request.form.get('student_id')
            conn = sqlite3.connect('database.db')
            conn.execute("DELETE FROM students WHERE id=?", (student_id,))
            conn.commit()
            conn.close()
            flash(f'🗑️ 学生 {student_id} 已删除', 'warning')
    
    conn = sqlite3.connect('database.db')
    students = conn.execute("SELECT * FROM students ORDER BY id").fetchall()
    conn.close()
    
    return render_template('manage_students.html', students=students)

@app.route('/admin/submissions')
@login_required
def view_submissions():
    conn = sqlite3.connect('database.db')
    submissions = conn.execute('''SELECT * FROM submissions 
        ORDER BY submitted_at DESC LIMIT 200''').fetchall()
    conn.close()
    return render_template('view_submissions.html', submissions=submissions)

@app.route('/admin/resolve_error/<int:error_id>', methods=['POST'])
@login_required
def resolve_error(error_id):
    resolution = request.form.get('resolution', '')
    correct_points = request.form.get('correct_points')
    
    if correct_points and correct_points.isdigit():
        conn = sqlite3.connect('database.db')
        conn.execute("""UPDATE submissions 
            SET final_points=?, status='resolved', reviewed_by=?, reviewed_at=CURRENT_TIMESTAMP
            WHERE id IN (SELECT submission_id FROM error_logs WHERE id=?)""",
                    (int(correct_points), current_user.id, error_id))
        conn.commit()
        conn.close()
    
    db.resolve_error(error_id, current_user.id, resolution)
    flash('✅ 错误已解决', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export_excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    ws_submissions = wb.active
    ws_submissions.title = "提交记录"
    headers = ['ID', '学号', '姓名', '班级', '用户输入', 'OCR识别', '最终积分', 
               '状态', '错误信息', '提交时间', '审核人', '审核时间']
    ws_submissions.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws_submissions[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    error_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    conn = sqlite3.connect('database.db')
    submissions = conn.execute('''SELECT * FROM submissions ORDER BY submitted_at DESC''').fetchall()
    
    for row in submissions:
        row_list = list(row)
        ws_submissions.append(row_list)
        
        last_row = ws_submissions.max_row
        if len(row) > 6 and row[6] == 'mismatch':
            for cell in ws_submissions[last_row]:
                cell.fill = warning_fill
        elif len(row) > 6 and row[6] == 'error':
            for cell in ws_submissions[last_row]:
                cell.fill = error_fill
    
    ws_stats = wb.create_sheet("统计信息")
    stats = db.get_statistics()
    ws_stats.append(['统计项', '数值'])
    for key, value in stats.items():
        ws_stats.append([key, value])
    
    ws_students = wb.create_sheet("学生列表")
    students = conn.execute("SELECT * FROM students ORDER BY id").fetchall()
    ws_students.append(['学号', '姓名', '班级', '添加时间'])
    for student in students:
        ws_students.append(list(student))
    
    conn.close()
    
    excel_path = f'student_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(excel_path)
    
    return send_file(excel_path, as_attachment=True, download_name='student_points_report.xlsx')

@app.route('/admin/api/pending_errors')
@login_required
def api_pending_errors():
    pending = db.get_pending_errors()
    return jsonify({'count': len(pending)})

# 创建必要的目录
os.makedirs('uploads', exist_ok=True)
os.makedirs('static', exist_ok=True)
